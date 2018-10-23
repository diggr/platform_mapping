#!/usr/bin/env python3
"""
Converts platform mapping tsv files to json, xlsx and csv files, to be used as a static rest api.

Copyright (C) 2018 Leipzig University Library <info@ub.uni-leipzig.de>

@author   Florian Rämisch <raemisch@ub.uni-leipzig.de>
@license  https://opensource.org/licenses/GPL-3.0 GNU GPLv3

This program is free software; you can redistribute it and/or modify
it under the terms of the GNU General Public License version 2,
as published by the Free Software Foundation.

 This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program; if not, write to the Free Software
Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  USA


"""
import csv
import datetime
import json
import io
import shutil
import markdown

from os.path import join, abspath, dirname, isdir
from os import mkdir, listdir
from openpyxl import Workbook
from jinja2 import Template

# CONFIGURATION

## DELIMITER OF THE BASE DATA
DELIMITER="\t"

## BASE PATH IS PREPENDED TO ALL OTHER DATA PATHS
BASE_PATH = join(dirname(abspath(__file__)), "..")

## SKELETON IS LOCATED AT
SKELETON_FILENAME = join( BASE_PATH, "skeleton.json")

## MAPPING DATA IS LOCATED AT
MAPPING_PATH = join(BASE_PATH, "tabular_data")
MAPPING_FILENAMES = { "ESRB" : "esrb.tsv",
                      "GameFAQs": "gamefaqs.tsv",
                      "MediaArtDB": "mediaartdb.tsv",
                      "Mobygames": "mobygames.tsv",
                      "OGDB": "ogdb.tsv",
                      "PEGI": "pegi.tsv",
                      "USK": "usk.tsv",
                      "Diggr Vocabulary": "diggr_vocab.tsv"}
MAPPING_FILES = { m_title:join(MAPPING_PATH, m_filename) for m_title, m_filename in MAPPING_FILENAMES.items() }

## PLATFORM GROUP FILE
PLATFORM_GROUP_FILE = join(MAPPING_PATH, "platform_groups.tsv")

## OUTPUT
OUT_PATH = join(BASE_PATH, "dist")
OUT_EXT = ".json"
PLATFORM_GROUP_OUTFILE = join(OUT_PATH, "platform_groups"+OUT_EXT)

## INDEX PAGE
INDEX_TEMPLATE = join(BASE_PATH, "templates", "index.html")
README = join(BASE_PATH, "README.md")
INDEX_PAGE = join(OUT_PATH, "index.html")

if not isdir(OUT_PATH):
    mkdir(OUT_PATH)

# MAPPING FUNCTIONS
def mapping_to_json(mapping_file, delimiter=INPUT_DELIMITER):
    """
    Opens a tabular (default) or any other kind of two-column csv-data
    and returns its content as python object
    """
    output = dict()
    with open(mapping_file) as mfile:
        reader = csv.reader(mfile, delimiter=delimiter)

        # skip header
        next(reader)

        for row in reader:
            output[row[0]] = row[1]
    return output


def mapping_to_xlsx(mapping_file, delimiter=DELIMITER):
    """
    Opens a tabular (default) or any other kind of two-column csv-data
    and returns its content as python object
    """
    with open(mapping_file) as mfile:
        reader = csv.reader(mfile, delimiter=delimiter)

        wb = Workbook()
        ws = wb.active

        for r, row in enumerate(reader,1):
            ws.cell(row=r, column=1, value=row[0])
            ws.cell(row=r, column=2, value=row[1])

    return wb


def mapping_to_csv(mapping_file, delimiter=DELIMITER):
    """
    Opens a tabular (default) or any other kind of two-column csv-data
    and returns its content as python object
    """
    with open(mapping_file) as mfile:
        reader = csv.reader(mfile, delimiter=delimiter)
        outfile = io.StringIO()
        output = csv.writer(outfile, delimiter=";", quotechar='"')

        for row in enumerate(reader):
            output.writerow(row[1])

        return outfile

def platform_group_to_json(platform_group_file, delimiter=DELIMITER):
    with open(platform_group_file) as pffile:
        reader = csv.reader(pffile, delimiter=delimiter)

        result = {}
        groupname = ""
        group = list()

        for row in reader:
            platform_name = row[1].strip()
            try:
                result[row[0]].append(platform_name)
            except KeyError:
                result[row[0]] = [platform_name]

        return result

def platform_group_to_xlsx(platform_group_file, delimiter=DELIMITER):
    with open(platform_group_file) as pffile:
        reader = csv.reader(pffile, delimiter=delimiter)

        wb = Workbook()
        ws = wb.active

        for r, row in enumerate(reader):
            ws.cell(row=1, column=0, value=row[0])
            ws.cell(row=1, column=1, value=row[1])

        return wb

def platform_group_to_csv(platform_group_file, delimiter=DELIMITER):
    with open(platform_group_file) as pffile:
        reader = csv.reader(pffile, delimiter=delimiter)

        outfile = io.StringIO()
        output = csv.writer(outfile, delimiter=";", quotechar='"')

        for r, row in enumerate(reader):
            output.writerow(row[1])

        return outfile

def load_skeleton(skeleton_filename=SKELETON_FILENAME):
    """
    Loads a json file and returns its content as
    python object.
    """
    with open(skeleton_filename) as sfile:
        return json.load(sfile)

def make_index_page(index_template=INDEX_TEMPLATE, readme=README, index_page=INDEX_PAGE, out_path=OUT_PATH):
    files = listdir(out_path)
    files.sort()
    for filename in files:
        ext = filename.split(".")[-1]
        print(filename, ext)
        if ext not in ["csv", "xlsx", "json"]:
            files.remove(filename)
    with open(readme) as readme_file:
        readme_text = readme_file.read()
    readme_html = markdown.markdown(readme_text)
    with open(index_template) as index:
        template = Template(index.read())
    with open(index_page, "w") as index_page_file:
        index_page_file.write(template.render(README=readme_html, files=files))

def create_static_rest_files(mapping_files=MAPPING_FILES,
                             platform_group_file=PLATFORM_GROUP_FILE,
                             platform_group_outfile=PLATFORM_GROUP_OUTFILE,
                             delimiter=DELIMITER,
                             out_path=OUT_PATH,
                             out_ext=OUT_EXT,
                             skeleton_filename=SKELETON_FILENAME):
    """
    Actual conversion function. Takes a list of mapping files and
    writes the json output files.
    """
    # Gather
    skeleton = load_skeleton(skeleton_filename)
    skeleton["meta"]["date"] = datetime.datetime.now().replace(microsecond=0).isoformat()

    for m_title, m_filename in mapping_files.items():
        print("Converting {} ...".format(m_title))
        json_mapping = mapping_to_json(m_filename)
        csv_mapping = mapping_to_csv(m_filename)
        xlsx_mapping = mapping_to_xlsx(m_filename)

        # Combine
        out_data = skeleton.copy()
        out_data["meta"]["name"] = m_title
        out_data["mapping"] = json_mapping

        # Save JSON
        json_out_file = join(out_path, m_title.lower().replace(" ", "_")+out_ext)
        with open(json_out_file, "w") as outfile:
            json.dump(out_data, outfile, indent=4)

        # Save CSV
        csv_out_file = join(out_path, m_title.lower().replace(" ", "_")+".csv")
        with open(csv_out_file, "w") as outfile:
            csv_mapping.seek(0)
            shutil.copyfileobj(csv_mapping, outfile)

        # Save XLSX
        xlsx_out_file = join(out_path, m_title.lower().replace(" ", "_")+".xlsx")
        xlsx_mapping.save(xlsx_out_file)

    #Platform Group
    print("Converting Platform Groups ...")
    platform_groups = platform_group_to_json(platform_group_file)
    pf_result = skeleton.copy()
    pf_result["name"] = "Platform Groups"
    pf_result["platform_groups"] = platform_groups
    with open(platform_group_outfile, "w") as outfile:
        json.dump(pf_result, outfile, indent=4)

if __name__ == "__main__":
    create_static_rest_files()
    make_index_page()
